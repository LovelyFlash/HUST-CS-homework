import json
import logging
import typing as t
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from fastapi import APIRouter, Request, Response, Depends, status, HTTPException
from sqlalchemy.orm import Session
from urllib.parse import quote
from sqlalchemy.sql import text
from sqlalchemy import distinct, and_, func, create_engine
from aioredis import Redis

from database.session import get_db
import models
import schemas
from schemas.basic import Response200, Response400
from database.redis import course_cache
from core.config import config
from .commonCourse import get_each_grade_class_number
import re

classDim_router = APIRouter()


@classDim_router.post("/scores/class/chart")
async def get_class_chart_by_term(
    queryItem: schemas.ClassQuery,
    db: Session = Depends(get_db),
    redis_store: Redis = Depends(course_cache),
):
    """
    班级维度图标

    :param queryItem:
    :param db:
    :param redis_store:
    :return: [
         {
            "term": "11",
            "grade": "18",
            "classNameList": [
                "卓越1801",
                "ACM1801",
                "计科1801",
            ],
            "failedNum": [
                0,
                0,
                3
            ],
            "failedRate": [
                0.0,
                0.0,
                11.0
            ]
         }
    ]
    """
    term = int(queryItem.term)
    key = "class_by_term_chart_" + str(term)
    sqlState = (
        db.query(models.ResultReadState)
        .filter(models.ResultReadState.key == key)
        .first()
    )
    redisState = await redis_store.exists(key)
    if (not sqlState) or redisState == 0:
        if not sqlState:  # 更新数据
            ret = []
            # 找到四个年级
            grade_class_info = await get_each_grade_class_number(db, redis_store)
            print("get_class_chart_by_term:{}".format(grade_class_info))
            for grade in grade_class_info["grade"].keys():
                tmpDict = {
                    "classNameList": [],
                    "failedNum": [],  # 挂科人数
                    "failedRate": [],
                    "totalNum": [],  # 各班总人数
                    "grade": str(grade),
                    "term": str(term),
                    "type": 1,
                }
                tmpDictMajor = {
                    "classNameList": [],
                    "failedNum": [],  # 挂科人数
                    "failedRate": [],
                    "totalNum": [],  # 各班总人数
                    "grade": str(grade),
                    "term": str(term),
                    "type": 0,
                }
                count = 0
                lastMajorIndex = 0
                for className in grade_class_info["class_nums_student"][grade]:
                    if (
                        count == 0
                        or re.search(
                            r"[A-Z]+", tmpDict["classNameList"][count - 1]
                        ).group()
                        != re.search(r"[A-Z]+", className).group()
                    ):
                        tmpDictMajor["classNameList"].append(
                            re.search(r"[A-Z]+", className).group()
                        )
                        tmpDictMajor["failedNum"].append(0)
                        tmpDictMajor["failedRate"].append(0)
                        tmpDictMajor["totalNum"].append(0)
                        lastMajorIndex += 1
                    res_class = (
                        db.query(models.Scores)
                        .join(
                            models.Students,
                            models.Scores.stuID == models.Students.stuID,
                            isouter=True,
                        )
                        .filter(
                            and_(
                                models.Students.stuClass == className,
                                models.Scores.term == term,
                                models.Scores.failed == 1,
                            )
                        )
                        .with_entities(models.Scores.stuID)
                        .distinct()
                        .all()
                    )
                    tmpDict["classNameList"].append(className)
                    tmpDict["failedNum"].append(len(res_class))
                    tmpDictMajor["failedNum"][lastMajorIndex - 1] += len(res_class)
                    tmpDict["failedRate"].append(0)
                    tmpDict["totalNum"].append(
                        grade_class_info["class_nums_student"][str(grade)][className]
                    )
                    tmpDictMajor["totalNum"][lastMajorIndex - 1] += grade_class_info[
                        "class_nums_student"
                    ][str(grade)][className]
                    count += 1

                # 计算不及格率，修改班级名
                for i in range(len(tmpDict["classNameList"])):
                    if tmpDict["totalNum"][i] != 0:
                        tmpDict["failedRate"][i] = round(
                            tmpDict["failedNum"][i] / tmpDict["totalNum"][i] * 100, 2
                        )

                    tmp_className = tmpDict["classNameList"][i]
                    if tmp_className == "":
                        continue
                    tmp_className = (
                        re.search(r"[A-Z]+", tmp_className).group()
                        + "\n"
                        + (
                            re.search(r"[0-9]+", tmp_className).group()
                            if re.search(r"[0-9]+", tmp_className) != None
                            else ""
                        )
                    )
                    tmp_className = tmp_className.replace("CS", "计科")
                    tmp_className = tmp_className.replace("BSB", "本硕博(启明)")
                    tmp_className = tmp_className.replace("BD", "大数据")
                    tmp_className = tmp_className.replace("IOT", "物联网")
                    tmp_className = tmp_className.replace("IST", "智能")
                    tmp_className = tmp_className.replace("ZY", "卓越(创新)")
                    tmp_className = tmp_className.replace("TL", "图灵")
                    tmp_className = tmp_className.replace("XJ", "校交")
                    tmpDict["classNameList"][i] = tmp_className
                if sum(tmpDict["failedNum"]) != 0:
                    ret.append(tmpDict)

                for i in range(len(tmpDictMajor["classNameList"])):
                    if tmpDictMajor["totalNum"][i] != 0:
                        tmpDictMajor["failedRate"][i] = round(
                            tmpDictMajor["failedNum"][i]
                            / tmpDictMajor["totalNum"][i]
                            * 100,
                            2,
                        )

                    tmp_className = tmpDictMajor["classNameList"][i]
                    if tmp_className == "":
                        continue
                    tmp_className = (
                        re.search(r"[A-Z]+", tmp_className).group()
                        + "\n"
                        + (
                            re.search(r"[0-9]+", tmp_className).group()
                            if re.search(r"[0-9]+", tmp_className) != None
                            else ""
                        )
                    )
                    tmp_className = tmp_className.replace("CS", "计科")
                    tmp_className = tmp_className.replace("BSB", "本硕博(启明)")
                    tmp_className = tmp_className.replace("BD", "大数据")
                    tmp_className = tmp_className.replace("IOT", "物联网")
                    tmp_className = tmp_className.replace("IST", "智能")
                    tmp_className = tmp_className.replace("ZY", "卓越(创新)")
                    tmp_className = tmp_className.replace("TL", "图灵")
                    tmp_className = tmp_className.replace("XJ", "校交")
                    tmpDictMajor["classNameList"][i] = tmp_className
                if sum(tmpDictMajor["failedNum"]) != 0:
                    ret.append(tmpDictMajor)

            if len(ret) != 0:
                # 将计算好的数据写入数据库
                for classItem in ret:
                    classItemQuery = (
                        db.query(models.ClassByTermChart)
                        .filter(
                            and_(
                                models.ClassByTermChart.term == classItem["term"],
                                models.ClassByTermChart.grade
                                == str(classItem["grade"]),
                                models.ClassByTermChart.type == classItem["type"],
                            )
                        )
                        .first()
                    )
                    if classItemQuery:
                        db.query(models.ClassByTermChart).filter(
                            and_(
                                models.ClassByTermChart.term == classItem["term"],
                                models.ClassByTermChart.grade
                                == str(classItem["grade"]),
                                models.ClassByTermChart.type == classItem["type"],
                            )
                        ).delete()
                        db.commit()
                    print("将 班级维度的表数据计算 结果存入数据库 ...")
                    classItemInsert = models.ClassByTermChart(
                        term=classItem["term"],
                        grade=str(classItem["grade"]),
                        classNameList=str(classItem["classNameList"]),
                        failedNum=str(classItem["failedNum"]),
                        failedRate=str(classItem["failedRate"]),
                        totalNum=str(classItem["totalNum"]),
                        type=classItem["type"],
                    )
                    db.add(classItemInsert)
                    db.commit()
                    db.refresh(classItemInsert)
                state_insert = models.ResultReadState(key=key)
                db.add(state_insert)
                db.commit()
                db.refresh(state_insert)

            else:
                return Response400(
                    msg="没有查询到相关数据，请检查查询关键字 or 数据库数据是否为空"
                )

        else:  # 直接读取已经计算好的数据
            retDb = (
                db.query(models.ClassByTermChart)
                .filter(models.ClassByTermChart.term == str(term))
                .all()
            )
            ret = []
            for dbItem in retDb:
                ret.append(
                    {
                        "term": dbItem.term,
                        "grade": dbItem.grade,
                        "classNameList": (
                            eval(dbItem.classNameList)
                            if dbItem.classNameList != ""
                            else []
                        ),
                        "failedNum": (
                            eval(dbItem.failedNum) if dbItem.classNameList != "" else []
                        ),
                        "failedRate": (
                            eval(dbItem.failedRate)
                            if dbItem.classNameList != ""
                            else []
                        ),
                        "totalNum": (
                            eval(dbItem.totalNum) if dbItem.totalNum != "" else []
                        ),
                        "type": dbItem.type,
                    }
                )
        if len(ret) != 0:
            # 将数据写入 redis
            await redis_store.setex(
                key,
                config.FAILED_STUID_INFO_REDIS_CACHE_EXPIRES,
                json.dumps(ret, ensure_ascii=False),
            )
        else:
            return Response400(
                data="中间结果数据库中暂时无数据，请在文件上传页面切换读取数据方式为从原始数据读取之后再请求"
            )
    else:
        ret = json.loads(await redis_store.get(key))

    return Response200(data=ret)


@classDim_router.post("/credits/class/export")
async def export_class_credits_detail(
    queryItem: schemas.ClassCreditExport,
    db: Session = Depends(get_db),
):
    """
    导出班级学分统计明细表格
    
    :param queryItem: 包含学期和班级名称的查询参数
    :param db: 数据库会话
    :return: Excel文件响应
    """
    term = int(queryItem.term)
    className = queryItem.className
    
    # 获取班级所有学生
    students = (
        db.query(models.Students)
        .filter(models.Students.stuClass == className)
        .order_by(models.Students.stuID)
        .all()
    )
    
    if not students:
        return Response400(msg=f"班级 {className} 没有学生数据")
    
    # 获取班级所有课程成绩
    student_ids = [student.stuID for student in students]
    scores = (
        db.query(models.Scores)
        .join(
            models.Courses,
            and_(
                models.Scores.courseName == models.Courses.courseName,
                models.Scores.term == models.Courses.term,
            ),
        )
        .filter(
            and_(
                models.Scores.stuID.in_(student_ids),
                models.Scores.term == term
            )
        )
        .with_entities(
            models.Scores.stuID,
            models.Scores.courseName,
            models.Scores.score,
            models.Scores.failed,
            models.Courses.credit,
            models.Courses.type
        )
        .all()
    )
    
    # 组织数据
    student_credits = {}
    course_details = {}
    
    for score in scores:
        stu_id = score.stuID
        course_name = score.courseName
        score_value = score.score
        failed = score.failed
        credit = score.credit or 0.0  # Handle None for credit
        course_type = score.type
        
        # 初始化学生数据
        if stu_id not in student_credits:
            student_credits[stu_id] = {
                'required_credits': 0.0,
                'specialized_elective_credits': 0.0,
                'failed_required_courses': [],
            }
        
        # 初始化课程详情
        if course_name not in course_details:
            course_details[course_name] = {
                'credit': credit,
                'type': course_type,
                'scores': {}
            }
        
        # 记录学生课程分数
        course_details[course_name]['scores'][stu_id] = score_value
        
        # 计算学分统计
        if failed == 0:  # 通过课程
            if course_type in [1, 2]:  # 公共必修或专业必修
                student_credits[stu_id]['required_credits'] += credit
            elif course_type == 3:  # 专业选修
                student_credits[stu_id]['specialized_elective_credits'] += credit
        else:  # 未通过课程
            if course_type in [1, 2]:  # 公共必修或专业必修
                student_credits[stu_id]['failed_required_courses'].append(course_name)    
    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = f"{className}学分统计"
    
    # 设置样式
    header_font = Font(bold=True, size=12)
    header_alignment = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 写入学分统计表头（第1-3行）
    ws['C1'] = '已获必修学分'
    ws['C2'] = '已获专选学分'
    ws['C3'] = '未通过必修课程'
    
    # 写入学生学分统计
    for i, student in enumerate(students, start=1):
        stu_id = student.stuID
        stats = student_credits.get(stu_id, {
            'required_credits': 0.0,
            'specialized_elective_credits': 0.0,
            'failed_required_courses': []
        })
        
        col_letter = get_column_letter(3 + i)  # D, E, F, ...
        ws[f'{col_letter}1'] = stats['required_credits']
        ws[f'{col_letter}2'] = stats['specialized_elective_credits']
        ws[f'{col_letter}3'] = "\n".join(stats['failed_required_courses'])
    
    # 课程详情从第4行开始
    course_header_row = 4
    
    # 写入课程详情表头
    ws[f'A{course_header_row}'] = '课程名'
    ws[f'B{course_header_row}'] = '课程类别'
    ws[f'C{course_header_row}'] = '课程学分'
    
    # 写入学生姓名
    for i, student in enumerate(students, start=1):
        col_letter = get_column_letter(3 + i)  # D, E, F, ...
        ws[f'{col_letter}{course_header_row}'] = student.stuName
    
    # 写入课程详情
    row = course_header_row + 1
    sorted_course_details = sorted(course_details.items())
    for course_name, course_info in sorted_course_details:
        ws[f'A{row}'] = course_name
        
        # 课程类型映射
        type_mapping = {
            0: '其他',
            1: '公共必修',
            2: '专业必修', 
            3: '专业选修',
            4: '公共选修',
            5: '重修课程'
        }
        ws[f'B{row}'] = type_mapping.get(course_info['type'], '其他')
        ws[f'C{row}'] = course_info['credit']
        
        # 写入每个学生的分数
        for i, student in enumerate(students, start=1):
            col_letter = get_column_letter(3 + i)  # D, E, F, ...
            score = course_info['scores'].get(student.stuID, '')
            ws[f'{col_letter}{row}'] = score
        
        row += 1
    
    # 应用样式
    max_col = 3 + len(students)
    # 学分统计部分
    for row_data in ws.iter_rows(min_row=1, max_row=3, max_col=max_col):
        for cell in row_data:
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border

    # 课程详情部分
    for row_data in ws.iter_rows(min_row=course_header_row, max_row=row - 1, max_col=max_col):
        for cell in row_data:
            cell.border = border
    
    # 调整列宽
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    
    for i in range(len(students)):
        col_letter = get_column_letter(4 + i)  # D, E, F, ...
        ws.column_dimensions[col_letter].width = 20 # Increased width for course names
        # Enable text wrapping for the failed courses row
        for r in ws.iter_rows(min_row=3, max_row=3, min_col=4, max_col=max_col):
            for cell in r:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

    
    # 保存到临时文件
    import tempfile
    import os
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    temp_file.close()
    
    # 读取文件内容并返回
    with open(temp_file.name, 'rb') as f:
        file_content = f.read()
    
    # 删除临时文件
    os.unlink(temp_file.name)
    
    from urllib.parse import quote

    # 返回Excel文件
    return Response(
        content=file_content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{quote(f'{className}_学分统计_{term}.xlsx')}"
        }
    )

@classDim_router.post("/credits/class/getjson")
async def export_class_credits_detail(queryItem: schemas.ClassCreditExport,db: Session = Depends(get_db),):
    """
    导出班级学分统计明细表格
    
    :param queryItem: 包含学期和班级名称的查询参数
    :param db: 数据库会话
    :return: json
    
    返回值样例:
    {
        "classInfo": [
            {
                "courseName": "高等数学",
                "credit": 4.0,
                "type": 1
            },
            {
                "courseName": "大学英语",
                "credit": 2.0,
                "type": 1
            }
        ],
        "students": [
            {
                "stuID": "20230001",
                "stuName": "张三",
                "required_credits": 6.0,
                "specialized_elective_credits": 0.0,
                "failed_required_courses": [],
                "stu_scores": [
                    {"course": "高等数学", "score": 90},
                    {"course": "大学英语", "score": 85}
                ]
            },
            {
                "stuID": "20230002",
                "stuName": "李四",
                "required_credits": 4.0,
                "specialized_elective_credits": 0.0,
                "failed_required_courses": ["大学英语"],
                "stu_scores": [
                    {"course": "高等数学", "score": 80},
                    {"course": "大学英语", "score": 55}
                ]
            }
        ]
    }
    
    """
    term = int(queryItem.term)
    className = queryItem.className
    
    # 获取班级所有学生
    students = (
        db.query(models.Students)
        .filter(models.Students.stuClass == className)
        .order_by(models.Students.stuID)
        .all()
    )
    
    if not students:
        return Response400(msg=f"班级 {className} 没有学生数据")
    
    abcded = db.query(models.Courses).all()
    for course in abcded:
        print(f"courseName: {course.courseName}, grade: {course.grade}, term: {course.term}, type: {course.type}")

    # 获取班级所有课程成绩
    student_ids = [student.stuID for student in students]
    scores = (
        db.query(models.Scores)
        .join(
            models.Courses,
            and_(
                models.Scores.courseName == models.Courses.courseName,
                models.Scores.term == models.Courses.term,
            ),
        )
        .filter(
            and_(
                models.Scores.stuID.in_(student_ids),
                models.Scores.term == term
            )
        )
        .with_entities(
            models.Scores.stuID,
            models.Scores.courseName,
            models.Scores.score,
            models.Scores.failed,
            models.Courses.credit,
            models.Courses.type
        )
        .all()
    )
    
    # 组织数据
    student_credits = {}
    course_details = {}
    course_names=[]
    
    for score in scores:
        stu_id = score.stuID
        course_name = score.courseName
        score_value = score.score
        failed = score.failed
        credit = score.credit or 0.0  # Handle None for credit
        course_type = score.type
        
        # 初始化学生数据
        if stu_id not in student_credits:
            student_credits[stu_id] = {
                'required_credits': 0.0,
                'specialized_elective_credits': 0.0,
                'failed_required_courses': [],
            }
        
        # 初始化课程详情
        if course_name not in course_details:
            course_details[course_name] = {
                'credit': credit,
                'type': course_type,
                'scores': {}
            }
            course_names.append(course_name)
        
        # 记录学生课程分数
        course_details[course_name]['scores'][stu_id] = score_value
        
        # 计算学分统计
        if failed == 0:  # 通过课程
            if course_type in [1, 2]:  # 公共必修或专业必修
                student_credits[stu_id]['required_credits'] += credit
            elif course_type == 3:  # 专业选修
                student_credits[stu_id]['specialized_elective_credits'] += credit
        else:  # 未通过课程
            if course_type in [1, 2]:  # 公共必修或专业必修
                student_credits[stu_id]['failed_required_courses'].append(course_name)
    
    # 存放主要信息
    class_info=[]
    for course_name in course_names:
        class_info.append({
            "courseName": course_name,
            "credit": course_details[course_name]['credit'],
            "type": course_details[course_name]['type'],
        })
    
    student_info = []
    for student in students:
        stu_scores=[]
        for i in course_names:
            if student.stuID in course_details[i]['scores']:
                stu_scores.append({
                    "course":i,
                    "score":course_details[i]['scores'][student.stuID]
                })
        student_info.append({
            "stuID": student.stuID,
            "stuName": student.stuName,
            "required_credits":student_credits[stu_id]['required_credits'],
            "specialized_elective_credits":student_credits[stu_id]['specialized_elective_credits'],
            "failed_required_courses":student_credits[stu_id]['failed_required_courses'],
            "stu_scores":stu_scores,
        })
    

    #创建json文件
    res={
        "classInfo":class_info,
        "students":student_info,
    }
    import tempfile
    import os
    import json
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.json', mode='w', encoding='utf-8')
    json.dump(res, temp_file, ensure_ascii=False, indent=4)
    temp_file.close()

    # 读取文件内容并返回
    with open(temp_file.name, 'r', encoding='utf-8') as f:
        file_content = f.read()

    # 删除临时文件
    os.unlink(temp_file.name)
    from urllib.parse import quote

    # 返回json文件
    return Response(
        content=file_content,
        media_type="application/json",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{quote(f'{className}_学分统计_{term}.json')}"
        }
    )